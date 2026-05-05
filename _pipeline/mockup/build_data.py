"""
Build mockup/data.js from the clean dashboard input.

Input:  sdwan_dashboard_input.xlsx (produced by build_input.py)
Output: mockup/data.js — window.SDWAN_DATA payload consumed by app.js

Thin transform: pivots prereqs long→nested, splits combined underlay/bandwidth fields,
maps the new 7-value status enum back to the legacy 4-value enum used by app.js
(keeps the rich enum in status_detail), and reconstructs risk_trend + prereq_meta.
"""

import argparse
import json
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "sdwan_dashboard_input.xlsx"
DEFAULT_OUTPUT = ROOT / "mockup" / "data.js"

# Map the rich 7-value enum to values the dashboard understands. "closed" is
# kept distinct from "migrated": Safran uses "closed" for sites taken out of
# the migration scope (decommissioned / dropped), NOT sites that were migrated
# to SD-WAN. Counting them as migrated over-estimates delivery.
LEGACY_STATUS_MAP = {
    "closed": "closed",
    "migrated": "migrated",
    "blocked": "to_validate",
    "ready": "ready",
    "planned": "to_validate",
    "postponed": "to_validate",
    "unknown": "unknown",
}

# Risk history bucket → human-readable label used by the dashboard trend chart
RISK_BUCKET_LABELS = {
    "sites_migrated": "Sites migrated",
    "no_risk": "no risk",
    "low_risk": "low risk",
    "medium_risk": "medium risk",
    "high_risk": "high risk",
    "not_possible_to_plan": "not possible to plan",
}


def read_sheet_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, r)) for r in rows[1:]]


def split_slash(s, idx):
    """Split a 'A / B' string and return element at idx, or None."""
    if not s:
        return None
    parts = [p.strip() for p in str(s).split(" / ")]
    if idx < len(parts) and parts[idx]:
        return parts[idx]
    return None


def baseline_week_year(iso_date):
    """From a YYYY-MM-DD baseline date, return (ISO week, year)."""
    if not iso_date:
        return None, None
    try:
        d = datetime.strptime(iso_date, "%Y-%m-%d").date()
        y, w, _ = d.isocalendar()
        return w, y
    except (ValueError, TypeError):
        return None, None


def main():
    parser = argparse.ArgumentParser(description="Build mockup/data.js from the clean input xlsx.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT),
                        help="Path to sdwan_dashboard_input.xlsx (default: project root)")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT),
                        help="Path to data.js (default: mockup/data.js)")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    wb = openpyxl.load_workbook(input_path, data_only=True)
    sites_raw = read_sheet_as_dicts(wb["sites"])
    prereqs_raw = read_sheet_as_dicts(wb["prereqs"])
    risk_raw = read_sheet_as_dicts(wb["risk_history"])
    rem_raw = read_sheet_as_dicts(wb["remediation"])

    # As-of date: read from README (row 4, col B)
    ws_readme = wb["0_README"]
    as_of = None
    for row in ws_readme.iter_rows(values_only=True):
        if row and row[0] == "As-of date":
            as_of = str(row[1])
            break
    as_of = as_of or datetime.now().strftime("%Y-%m-%d")

    # --- Pivot prereqs: site_id → {prereq_key: value_pct} ---
    prereqs_by_site = defaultdict(dict)
    prereq_meta = {}
    for p in prereqs_raw:
        prereqs_by_site[p["site_id"]][p["prereq_key"]] = p["value_pct"]
        key = p["prereq_key"]
        if key not in prereq_meta:
            prereq_meta[key] = {
                "when_days": p["when_days"],
                "when_label": p["when_label"],
                "owner": p["owner"],
            }

    # --- Index remediation by site_id (for inlining into site records) ---
    rem_by_site = {}
    for r in rem_raw:
        sid = r.get("site_id")
        if sid:
            rem_by_site[sid] = r

    # --- Build site records in the shape app.js expects ---
    sites = []
    for r in sites_raw:
        sid = r["site_id"]
        new_status = r["status"] or "unknown"
        legacy_status = LEGACY_STATUS_MAP.get(new_status, "unknown")
        bw_iso = r.get("migration_date_baseline")
        bw, by = baseline_week_year(bw_iso)

        rem = rem_by_site.get(sid, {})

        sites.append({
            "site_id": sid,
            "company": r["company"],
            "topology": r["topology"],
            "batch": r["batch_year"],
            "batch_underlay": r["region"],
            "address": r["address"],
            "it_contact": r["it_contact"],
            "fac_contact": r["facilities_contact"],
            "status_raw": r["status_raw"],
            # Dashboard-compatible status (legacy 4-value enum)
            "status": legacy_status,
            # Rich enum + flags (new)
            "status_detail": new_status,
            "status_inferred": bool(r["status_inferred"]),
            "partial_migration": bool(r["partial_migration"]),
            "remediation_open": bool(r["remediation_open"]),
            "migration_date": r["migration_date"],
            "migration_date_baseline": bw_iso,
            "baseline_week": bw,
            "baseline_year": by,
            "t_minus_days": r["t_minus_days"],
            "go_nogo_date": r["go_nogo_date"],
            "chrono_review_date": r["chrono_review_date"],
            "underlay_asis": r["underlay_asis"],
            "underlay_target_primary": split_slash(r["underlay_target"], 0),
            "underlay_target_secondary": split_slash(r["underlay_target"], 1),
            "bandwidth_primary": split_slash(r["bandwidth"], 0),
            "bandwidth_secondary": split_slash(r["bandwidth"], 1),
            "hypercare": r["hypercare"],
            "decom_change": r["decom_change"],
            "decom_status": r["decom_status"],
            "decom_date": r["decom_date"],
            "migration_referent": r["referent"],
            "coordination": r["coordinator"],
            "comment": r["comment"],
            "prereqs": prereqs_by_site.get(sid, {}),
            "risk_level": r["risk_level"],
            "risk_underlay": r["risk_underlay"],
            "risk_overlay": r["risk_overlay"],
            "risk_safran": r["risk_safran"],
            "risk_other": r["risk_other"],
            "remediation_cause": rem.get("root_cause"),
            "remediation_blocker": rem.get("blocker"),
            "remediation_action": rem.get("action_log"),
            "remediation_severity": rem.get("severity"),
            "remediation_category": rem.get("category"),
            "source": (r["source_sheets"] or "").split(", ") if r["source_sheets"] else [],
            # Legacy fields kept as None for compatibility (no longer populated)
            "compliance_pct": None,
            "compliance_details": None,
        })

    # --- Risk trend: label → {week: count} ---
    risk_trend = defaultdict(dict)
    for r in risk_raw:
        label = RISK_BUCKET_LABELS.get(r["bucket"], r["bucket"])
        if r["count"] is not None:
            risk_trend[label][r["week_iso"]] = r["count"]
    risk_trend = {k: v for k, v in risk_trend.items()}

    # --- Final payload ---
    out = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "as_of_date": as_of,
        "sites": sites,
        "prereq_meta": prereq_meta,
        "risk_trend": risk_trend,
    }

    payload = json.dumps(out, ensure_ascii=False, default=str)
    output_path.write_text(f"window.SDWAN_DATA = {payload};\n", encoding="utf-8")

    # --- Sanity check ---
    from collections import Counter
    print(f"wrote {output_path}")
    print(f"  sites: {len(sites)}")
    print(f"  legacy status: {dict(Counter(s['status'] for s in sites))}")
    print(f"  status_detail: {dict(Counter(s['status_detail'] for s in sites))}")
    print(f"  prereq_meta: {len(prereq_meta)} keys")
    print(f"  risk_trend: {len(risk_trend)} buckets, weeks: {sorted({w for v in risk_trend.values() for w in v})}")
    print(f"  with remediation info: {sum(1 for s in sites if s['remediation_open'])}")


if __name__ == "__main__":
    main()
