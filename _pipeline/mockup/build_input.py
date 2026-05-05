"""
Build sdwan_dashboard_input.xlsx — the single canonical input for the dashboard.

Reads:
  - SDWAN_Timeline_Intern_20240422.xlsx
  - Sprint SD-WAN Week - RAMPUP_20240422.xlsx

Writes:
  - sdwan_dashboard_input.xlsx (8 sheets)

Status enum (priority when multiple conditions match):
  closed > migrated > blocked > ready > planned > postponed > unknown
"""

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from build_input_orange import extract_orange_prereqs, OB_PREREQS

ROOT = Path(__file__).resolve().parent.parent


def latest_file(pattern):
    """Return the most recently modified file matching the glob pattern, or None."""
    candidates = sorted(ROOT.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


# Auto-detected defaults (latest dated file); --intern / --sprint override these.
def latest_dated_file(name_root):
    """Find the dated xlsx file with the most recent date in its filename.

    Accepts both `..._YYYYMMDD.xlsx` (underscore) and `... YYYYMMDD.xlsx` (space)
    naming conventions. Falls back to the most recently modified file.
    """
    import re
    candidates = list(ROOT.glob(f"{name_root}*.xlsx"))
    candidates = [c for c in candidates if not c.name.startswith("~$")]
    if not candidates:
        return None
    # Prefer files with a YYYYMMDD embedded in the name, sorted by that date desc.
    dated = []
    for c in candidates:
        m = re.search(r"(\d{8})", c.name)
        if m:
            dated.append((m.group(1), c))
    if dated:
        dated.sort(key=lambda x: x[0], reverse=True)
        return dated[0][1]
    # Otherwise fall back to mtime
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


DEFAULT_INTERN = latest_dated_file("SDWAN_Timeline_Intern") or ROOT / "SDWAN_Timeline_Intern.xlsx"
DEFAULT_SPRINT = latest_dated_file("Sprint SD-WAN Week - RAMPUP") or ROOT / "Sprint SD-WAN Week - RAMPUP.xlsx"
DEFAULT_OUTPUT = ROOT / "sdwan_dashboard_input.xlsx"

# Overridable at runtime via --as-of
TODAY = date.today()

# Known Site ID typos/variants: map variant -> canonical form.
# Empty in this public skeleton — populate locally if your extracts have
# recurring typos to normalize. Each entry: { variant_form: canonical_form }.
SITE_ID_ALIASES = {}

# Cells that look like garbage / header leaks and should be rejected.
# Any other Site ID value from BATCH/PILOTS is accepted as-is, even if
# non-standard (e.g. "SAB - Capaul 1", "SSA Hyderabad", "FRXXXX-1-S006-EVRY")
# — these are real sites the PMO team tracks with placeholder names awaiting
# a proper ID.
SITE_ID_JUNK_EXACT = {"1", "SITE ID", "SITE", "WHO", "WHEN", "COMPANY ID"}

# Explicit exclusions: sites forcefully dropped from the master even if
# present in BATCH / PILOTS. Empty by default in this public skeleton.
SITE_ID_EXCLUDE = set()

# 1-indexed column numbers in BATCH 25-26 (header row = row 5, data starts row 6).
# Keys MUST match mockup/app.js CATEGORIES[].prereqs entries — do not rename without
# updating app.js too.
PREREQ_COLS = {
    "Companies Validation": 18,
    "Contact": 19,
    "Underlay Req": 20,
    "Underlay Quote": 21,
    "Underlay Order": 22,
    "Overlay Req": 23,
    "Overlay Quote": 24,
    "Overlay Order": 25,
    "Site Constraints": 26,
    "Site Survey Overlay": 27,
    "Site Survey Underlay": 28,
    "LLD As Is": 29,
    "Data Collect File": 30,
    "FOF Change": 31,
    "LLD Ready": 32,
    "Chronogram Ready": 33,
    "Dynatrace": 34,
    "Remediation": 35,
    "In-House Cabling": 36,
    "CMDB Change": 37,
    "Underlay Delivered": 38,
    "Overlay Delivered": 39,
    "LAN Switch Ready": 40,
    "C-Edge Install": 41,
    "CSR Request": 42,
    "Migration Change": 43,
    "BGP Preparation": 44,
    "Chrono Review": 45,
}

# When the source row 3 (WHEN) is empty, fall back to these defaults (days, negative = before T0).
DEFAULT_WHEN_DAYS = {
    "Companies Validation": -98,
    "Contact": -91,
    "Underlay Req": -84,
    "Underlay Quote": -77,
    "Underlay Order": -70,
    "Overlay Req": -84,
    "Overlay Quote": -77,
    "Overlay Order": -70,
    "Site Constraints": -63,
}

# Normalization rules (printed in README sheet)
NORMALIZATION_RULES = [
    "Canonical scope = UNION(BATCH 25-26, PILOTS) only — no other sheet defines scope",
    "Other sheets (Decomm+SNOW, Risk follow-up, etc.) only ENRICH existing sites — never add new ones",
    "All non-empty Site ID values from BATCH/PILOTS are kept — non-standard formats accepted",
    "Only stripped: header leaks (Site ID, WHO, WHEN, Company ID) and the literal value '1'",
    "Site IDs canonicalized via SITE_ID_ALIASES (configurable in build_input.py)",
    "Explicit exclusions configurable in SITE_ID_EXCLUDE (build_input.py)",
    "",
    "Status enum (priority order): closed > migrated > blocked > ready > planned > postponed > unknown",
    "  · 'Partial Migration' → migrated (with partial_migration=True flag)",
    "  · 'To validate' + no date → postponed",
    "  · Site in active Remediation → blocked (overrides raw status)",
    "  · No status_raw + past date → migrated (status_inferred=True)",
    "  · No status_raw + future date → planned (status_inferred=True)",
    "  · No status_raw + no date → postponed (status_inferred=True)",
    "",
    "Dates: ISO format YYYY-MM-DD (no time component)",
    "Prereq values: 0/1/50/100 (percent-like ints); NA kept as string in value_raw",
    "Prereq due_date = migration_date + when_days; overdue = today past due AND value < 100",
    "Weekly plan: rows 2-10 of Planif only; rows 11+ treated as legend",
    "Weekly plan fuzzy match: longest-prefix match on normalized short name (min 6 chars)",
    "Decom backlog: overdue = migrated > 60 days ago AND no CHG number",
]


def canon_key(s):
    if s is None:
        return ""
    s = str(s).upper().strip()
    return re.sub(r"\s+", "", s)


def norm_key(s):
    if s is None:
        return ""
    s = str(s).upper()
    return re.sub(r"[^A-Z0-9]", "", s)


def to_iso(v):
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        return d.strftime("%Y-%m-%d")
    return None


def norm_owner(raw):
    if not raw:
        return "Shared"
    s = str(raw).upper()
    saf = any(t in s for t in ("COMPANIES", "DSIC", "SAFRAN", "SAF", "DIGIT"))
    obs = any(t in s for t in ("OBS", "AT&T"))
    if saf and obs:
        return "Shared"
    if saf:
        return "Safran"
    if obs:
        return "OBS"
    return "Shared"


def parse_when(v):
    if v is None or v == "":
        return None
    s = str(v).strip().upper().replace(" ", "")
    m = re.match(r"([+-]?)(\d+)([WD])", s)
    if not m:
        if s in ("DD", "/"):
            return 0
        return None
    sign = -1 if m.group(1) == "-" else 1
    n = int(m.group(2))
    unit = m.group(3)
    return sign * n * (7 if unit == "W" else 1)


def derive_business_unit(site_id):
    # CC<num>-<batch>-<BU>-<location> → BU
    parts = str(site_id).split("-")
    return parts[2] if len(parts) >= 4 else None


def derive_country(site_id):
    # CC<num>-<batch>-<BU>-<location> → CC
    s = str(site_id)
    return s[:2] if len(s) >= 2 else None


def classify_status(status_raw, migration_date, has_remediation):
    """Returns (status, inferred_bool). Inferred=True when falling back to date-based guess."""
    s = (status_raw or "").strip().lower()
    if s == "closed":
        return "closed", False
    if s == "migrated" or s == "migrated / closed":
        return "migrated", False
    if s == "partial migration":
        # Sites marked "Partial Migration" with a future migration_date are
        # really still planning (waves split across dates) — not yet done.
        # Only treat as migrated when the migration date is in the past.
        if migration_date:
            try:
                md = datetime.strptime(migration_date, "%Y-%m-%d").date()
                if md > TODAY:
                    return "planned", False
            except (ValueError, TypeError):
                pass
        return "migrated", False
    if has_remediation:
        return "blocked", False
    if s == "ready":
        return "ready", False
    if s == "to validate":
        return "planned" if migration_date else "postponed", False
    # No status_raw: fall back to migration_date
    if migration_date is None:
        return "postponed", True
    try:
        md = datetime.strptime(migration_date, "%Y-%m-%d").date()
        return ("migrated" if md < TODAY else "planned"), True
    except (ValueError, TypeError):
        return "postponed", True


def _apply_alias(site_id):
    return SITE_ID_ALIASES.get(site_id, site_id)


def load_canonical_sites(wb_i, dupes_log, invalid_log, excluded_log):
    """Build canonical master = UNION of BATCH 25-26 + PILOTS (aliases applied).
    Per PMO arbitrage: these two sheets are the ONLY authoritative lists of
    sites in the migration scope. Any site that should be in the dashboard
    must be present here; conversely 'Sites list', 'Decom', 'Risk follow-up',
    etc. are NOT used for scope definition — they only enrich existing sites.
    """
    sites = {}
    by_canon = {}

    def try_add(site_raw, company_raw, mig, source_label, row_idx, extras=None):
        if site_raw is None or site_raw == "":
            return
        site = _apply_alias(str(site_raw).strip())
        if not site or site.upper() in SITE_ID_JUNK_EXACT:
            invalid_log.append((source_label, row_idx, str(site_raw).strip()))
            return
        if site in SITE_ID_EXCLUDE:
            excluded_log.append((source_label, row_idx, site))
            return
        ck = canon_key(site)
        company = str(company_raw).strip() if company_raw else None
        if ck in by_canon:
            existing = by_canon[ck]
            # Same canonical ID already added from a previous sheet — just
            # track the additional source.
            if source_label not in sites[existing]["source_sheets"]:
                sites[existing]["source_sheets"].append(source_label)
            dupes_log.append((site, existing, row_idx))
            return
        sites[site] = _base_record(site, company, mig)
        sites[site]["source_sheets"] = [source_label]
        if extras:
            for k, v in extras.items():
                if v is not None and sites[site].get(k) in (None, ""):
                    sites[site][k] = v
        by_canon[ck] = site

    # ------ Pass 1: BATCH 25-26 ------
    ws_b = wb_i["BATCH 25-26"]
    for i in range(6, ws_b.max_row + 1):
        raw = ws_b.cell(row=i, column=1).value
        company = ws_b.cell(row=i, column=4).value
        mig = ws_b.cell(row=i, column=47).value
        try_add(raw, company, mig, "BATCH 25-26", i)

    # ------ Pass 2: PILOTS ------
    # Header row 11; data from row 12. Cols: 1=Site ID, 2=Site, 3=Company,
    # 4=Address, 5=IT Contact, 7=Topology, 8=Underlay ASIS, 9=Underlay Target,
    # 11=Status (e.g. "Migrated", "Ready").
    ws_p = wb_i["PILOTS"]
    for i in range(12, ws_p.max_row + 1):
        raw = ws_p.cell(row=i, column=1).value
        company = ws_p.cell(row=i, column=3).value
        extras = {
            "address":      ws_p.cell(row=i, column=4).value,
            "it_contact":   ws_p.cell(row=i, column=5).value,
            "topology":     ws_p.cell(row=i, column=7).value,
            "underlay_asis":ws_p.cell(row=i, column=8).value,
            "underlay_target": ws_p.cell(row=i, column=9).value,
            "status_raw":   ws_p.cell(row=i, column=11).value,
        }
        # Normalize string fields
        for k in ("address", "it_contact", "topology", "underlay_asis", "underlay_target", "status_raw"):
            if extras.get(k) is not None:
                extras[k] = str(extras[k]).strip() or None
        try_add(raw, company, None, "PILOTS", i, extras=extras)

    return sites, by_canon


def _base_record(site, company, mig):
    return {
            "site_id": site,
            "company": company,
            "business_unit": derive_business_unit(site),
            "country": derive_country(site),
            "region": None,
            "topology": None,
            "batch_year": None,
            "address": None,
            "it_contact": None,
            "facilities_contact": None,
            "status": "unknown",
            "status_raw": None,
            "status_inferred": False,
            "partial_migration": False,
            "migration_date": to_iso(mig),
            "migration_date_baseline": None,
            "go_nogo_date": None,
            "chrono_review_date": None,
            "hypercare": None,
            "underlay_asis": None,
            "underlay_target": None,
            "bandwidth": None,
            "prereq_progress_pct": None,
            "risk_level": None,
            "risk_underlay": None,
            "risk_overlay": None,
            "risk_safran": None,
            "risk_other": None,
            "remediation_open": False,
            "decom_change": None,
            "decom_status": None,
            "decom_date": None,
            "referent": None,
            "coordinator": None,
            "comment": None,
            "t_minus_days": None,
            "source_sheets": [],
    }


def make_resolver(by_canon, orphans_log):
    def resolve(raw, source):
        if not raw:
            return None
        s = str(raw).strip()
        if s in SITE_ID_ALIASES:
            s = SITE_ID_ALIASES[s]
        resolved = by_canon.get(canon_key(s))
        if resolved is None:
            orphans_log[source].append(s)
        return resolved
    return resolve


def enrich_from_batch(sites, resolve, ws):
    """BATCH 25-26: status, dates, prereqs, topology, contacts. Also extract prereq_meta."""
    prereq_meta = {}
    for key, col in PREREQ_COLS.items():
        raw_when = ws.cell(row=3, column=col).value
        raw_who = ws.cell(row=4, column=col).value
        when_days = parse_when(raw_when)
        if when_days is None:
            when_days = DEFAULT_WHEN_DAYS.get(key, -7)
        prereq_meta[key] = {
            "when_days": when_days,
            "when_label": f"T{when_days // 7}W" if when_days % 7 == 0 and when_days != 0 else (
                "T0" if when_days == 0 else f"T{when_days}D"),
            "owner": norm_owner(raw_who),
            "owner_raw": str(raw_who).strip() if raw_who else None,
        }

    prereqs_long = []
    for i in range(6, ws.max_row + 1):
        raw_site = ws.cell(row=i, column=1).value
        if not raw_site:
            continue
        canon = resolve(raw_site, "BATCH 25-26")
        if canon is None:
            continue
        rec = sites[canon]

        rec["topology"] = ws.cell(row=i, column=2).value or rec["topology"]
        rec["region"] = ws.cell(row=i, column=3).value or rec["region"]
        rec["address"] = ws.cell(row=i, column=5).value or rec["address"]
        rec["it_contact"] = ws.cell(row=i, column=6).value or rec["it_contact"]
        rec["facilities_contact"] = ws.cell(row=i, column=7).value or rec["facilities_contact"]
        batch_year = ws.cell(row=i, column=8).value
        rec["batch_year"] = str(batch_year) if batch_year else rec["batch_year"]

        rec["underlay_asis"] = ws.cell(row=i, column=10).value or rec["underlay_asis"]
        under1 = ws.cell(row=i, column=11).value
        under2 = ws.cell(row=i, column=12).value
        rec["underlay_target"] = " / ".join(str(x) for x in [under1, under2] if x) or rec["underlay_target"]
        bw1 = ws.cell(row=i, column=13).value
        bw2 = ws.cell(row=i, column=14).value
        rec["bandwidth"] = " / ".join(str(x) for x in [bw1, bw2] if x) or rec["bandwidth"]

        prog = ws.cell(row=i, column=17).value
        try:
            rec["prereq_progress_pct"] = float(prog) * 100 if prog is not None else None
        except (ValueError, TypeError):
            rec["prereq_progress_pct"] = None

        # Prereqs long format
        for key, col in PREREQ_COLS.items():
            v = ws.cell(row=i, column=col).value
            value_pct = None
            value_raw = None
            if isinstance(v, (int, float)):
                value_pct = float(v)
            elif v is not None:
                value_raw = str(v).strip()
            when_days = prereq_meta[key]["when_days"]
            due_date = None
            overdue = None
            # Use the upcoming migration date if available
            mig_raw = ws.cell(row=i, column=47).value
            if isinstance(mig_raw, (datetime, date)):
                md = mig_raw.date() if isinstance(mig_raw, datetime) else mig_raw
                due_date = (md + (date.fromordinal(md.toordinal() + when_days) - md)).strftime("%Y-%m-%d") if when_days else md.strftime("%Y-%m-%d")
                # Overdue = today past due AND value < 100
                due = date.fromordinal(md.toordinal() + when_days)
                overdue = (TODAY >= due) and (value_pct is not None and value_pct < 100)
            prereqs_long.append({
                "site_id": canon,
                "prereq_key": key,
                "when_days": when_days,
                "when_label": prereq_meta[key]["when_label"],
                "owner": prereq_meta[key]["owner"],
                "value_pct": value_pct,
                "value_raw": value_raw,
                "due_date": due_date,
                "overdue": overdue,
            })

        rec["chrono_review_date"] = to_iso(ws.cell(row=i, column=45).value)
        rec["go_nogo_date"] = to_iso(ws.cell(row=i, column=46).value)
        mig_date = ws.cell(row=i, column=47).value
        if isinstance(mig_date, (datetime, date)):
            rec["migration_date"] = to_iso(mig_date)
            md = mig_date.date() if isinstance(mig_date, datetime) else mig_date
            rec["t_minus_days"] = (md - TODAY).days
        elif mig_date is not None:
            # "TBC" or other string - leave migration_date from Sites list, no t_minus
            pass

        status_raw = ws.cell(row=i, column=48).value
        rec["status_raw"] = str(status_raw).strip() if status_raw is not None else rec["status_raw"]
        rec["partial_migration"] = (str(status_raw or "").strip().lower() == "partial migration")

        rec["referent"] = ws.cell(row=i, column=49).value or rec["referent"]
        rec["coordinator"] = ws.cell(row=i, column=50).value or rec["coordinator"]
        rec["hypercare"] = ws.cell(row=i, column=51).value or rec["hypercare"]
        rec["comment"] = ws.cell(row=i, column=54).value or rec["comment"]

        if "BATCH 25-26" not in rec["source_sheets"]:
            rec["source_sheets"].append("BATCH 25-26")

    return prereq_meta, prereqs_long


def enrich_from_risk(sites, resolve, ws):
    """Risk follow-up: risk level + subcategories."""
    for i in range(13, ws.max_row + 1):
        raw_site = ws.cell(row=i, column=1).value
        if not raw_site:
            continue
        canon = resolve(raw_site, "Risk follow-up")
        if canon is None:
            continue
        rec = sites[canon]
        rec["risk_level"] = (ws.cell(row=i, column=2).value or "").strip() or None
        rec["risk_underlay"] = (ws.cell(row=i, column=3).value or "").strip() or None
        rec["risk_overlay"] = (ws.cell(row=i, column=5).value or "").strip() or None
        rec["risk_safran"] = (ws.cell(row=i, column=6).value or "").strip() or None
        rec["risk_other"] = (ws.cell(row=i, column=7).value or "").strip() or None
        if "Risk follow-up" not in rec["source_sheets"]:
            rec["source_sheets"].append("Risk follow-up")


def enrich_from_miseenforme(sites, resolve, ws):
    """Mise en forme: baseline week/year for OTD."""
    for i in range(3, ws.max_row + 1):
        raw_site = ws.cell(row=i, column=5).value
        if not raw_site:
            continue
        canon = resolve(raw_site, "Mise en forme")
        if canon is None:
            continue
        bw = ws.cell(row=i, column=13).value
        by = ws.cell(row=i, column=14).value
        if isinstance(bw, (int, float)) and isinstance(by, (int, float)):
            # ISO week -> date (Monday of that week)
            try:
                d = date.fromisocalendar(int(by), int(bw), 1)
                sites[canon]["migration_date_baseline"] = d.strftime("%Y-%m-%d")
                if "Mise en forme" not in sites[canon]["source_sheets"]:
                    sites[canon]["source_sheets"].append("Mise en forme")
            except ValueError:
                pass


def enrich_from_decomm(sites, resolve, ws):
    """Decomm + SNOW + Dynatrace: decom change number + status."""
    # Header on row 1; Col 2 (1-indexed) = Site (full site_id), Col 16 = Migration Status,
    # Col 17 = Migration Date, Col 21 = Change DECOM N°
    for i in range(3, ws.max_row + 1):
        raw_site = ws.cell(row=i, column=2).value
        if not raw_site:
            continue
        canon = resolve(raw_site, "Decomm + SNOW + Dynatrace")
        if canon is None:
            continue
        rec = sites[canon]
        chg = ws.cell(row=i, column=21).value
        decom_status_raw = ws.cell(row=i, column=16).value
        decom_date = ws.cell(row=i, column=17).value
        if chg:
            rec["decom_change"] = str(chg).strip()
        if decom_status_raw:
            rec["decom_status"] = str(decom_status_raw).strip()
        if isinstance(decom_date, (datetime, date)):
            rec["decom_date"] = to_iso(decom_date)
        if "Decomm + SNOW + Dynatrace" not in rec["source_sheets"]:
            rec["source_sheets"].append("Decomm + SNOW + Dynatrace")


def extract_remediation(ws, sites):
    """Remédiation sheet → list of tickets with fuzzy site_id match."""
    # Build reverse index for fuzzy match
    site_norm_index = {}
    for sid in sites:
        site_norm_index[norm_key(sid)] = sid
        parts = sid.split("-")
        if len(parts) >= 4:
            short = f"{parts[2]} {parts[3]}"
            site_norm_index.setdefault(norm_key(short), sid)

    entries = []
    current_severity = "medium"
    for i in range(4, ws.max_row + 1):
        label = ws.cell(row=i, column=2).value
        cause = ws.cell(row=i, column=3).value
        blocker = ws.cell(row=i, column=4).value
        action = ws.cell(row=i, column=5).value
        if not label:
            continue
        label = str(label).strip()
        if label == "High Risks":
            current_severity = "high"
            continue
        if label == "Site":
            current_severity = "medium"
            continue

        # Fuzzy match
        key = norm_key(label)
        matched = None
        for k, sid in site_norm_index.items():
            if not k:
                continue
            if k in key or (len(k) >= 6 and key.startswith(k[:6])):
                matched = sid
                break
        # Infer category from root cause
        cause_s = str(cause or "").lower()
        if "underlay" in cause_s:
            category = "Underlay"
        elif "overlay" in cause_s:
            category = "Overlay"
        elif "construction" in cause_s or "travaux" in cause_s:
            category = "Site works"
        elif cause_s in ("", "n/a", "na"):
            category = "N/A"
        else:
            category = "Other"

        entries.append({
            "display_name": label,
            "site_id": matched,
            "category": category,
            "severity": current_severity,
            "root_cause": str(cause) if cause else None,
            "blocker": str(blocker) if blocker else None,
            "action_log": str(action) if action else None,
        })

    # Flag sites with open remediation
    for e in entries:
        if e["site_id"] and e["site_id"] in sites:
            sites[e["site_id"]]["remediation_open"] = True

    return entries


LEGEND_TOKENS = (
    "migration confirmée", "not secured", "high risks", "underlay", "overlay",
    "new overlay", "n/a mix", "pending sae", "pending", "root cause",
    "postponed or to be included", "stagging by safran", "sep (travaux)",
    "sae (pending)", "w39 à w18", "w21 à w18", "postponed ---> secure ok",
    "retrofit swap", "vmanage resiliency", "secure", "site",
)

# Explicit aliases for short names used in weekly plan that can't be auto-resolved
# (normalized key -> Site ID in master). Empty by default; populate locally as needed.
WEEKLY_PLAN_ALIASES = {}


def _clean_display_name(text):
    """Strip common suffixes so fuzzy matching can find the site."""
    s = text
    # Remove parenthesized qualifiers
    s = re.sub(r"\([^)]*\)", " ", s)
    # Remove dates like 15/04, 14/04/2026
    s = re.sub(r"\d{1,2}/\d{1,2}(?:/\d{2,4})?", " ", s)
    # Remove times like 10H, 9H30, 06AM, 18:00, 6:00 PM, CET
    s = re.sub(r"\d{1,2}[hH:][\dh:]*\s*(?:AM|PM|CET)?", " ", s)
    s = re.sub(r"\b(?:AM|PM|CET|HNO|WE|HO)\b", " ", s, flags=re.IGNORECASE)
    # Remove batch markers
    s = re.sub(r"\b(?:20\d\d[-]?[12]?|2Q\d\d|4Q\d\d)\b", " ", s)
    # Remove "by OB", "by SAF", etc.
    s = re.sub(r"\bby\s+\w+\b", " ", s, flags=re.IGNORECASE)
    return s


def extract_weekly_plan(ws, sites):
    """Planif migrations per week → long-format rows with fuzzy site_id match.
    Only rows 2-10 are actual weekly entries; rows 11+ are legend/meta."""
    # Build fuzzy index: map every (normalized) short form to a canonical site_id.
    # Short forms: full site_id, "BU CITY", "CITY", "BU CITYPREFIX"
    site_norm_index = {}
    for sid in sites:
        site_norm_index[norm_key(sid)] = sid
        parts = sid.split("-")
        if len(parts) >= 4:
            bu = parts[2]
            city = parts[3]
            site_norm_index.setdefault(norm_key(f"{bu} {city}"), sid)
            site_norm_index.setdefault(norm_key(city), sid)

    # Sort keys by length desc so we match the most specific first
    sorted_keys = sorted(site_norm_index.keys(), key=len, reverse=True)

    rows = []
    week_labels = {}
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=j).value
        if v:
            week_labels[j] = str(v).strip()

    MAX_PLAN_ROW = 10  # Rows 11+ are legend/meta

    for j, label in week_labels.items():
        for i in range(2, MAX_PLAN_ROW + 1):
            cell = ws.cell(row=i, column=j).value
            if not cell:
                continue
            text = str(cell).strip()
            if not text:
                continue
            t_low = text.lower()
            # Skip legend/meta tokens
            if any(tok in t_low for tok in LEGEND_TOKENS):
                # SPARE entries should still be kept (they're reserved slots)
                if "spare" not in t_low:
                    continue

            # Classify confidence
            if "spare" in t_low:
                confidence = "spare"
            elif "tbc" in t_low or "best effort" in t_low:
                confidence = "placeholder"
            elif "rollback" in t_low or "postpone" in t_low:
                confidence = "failed"
            else:
                confidence = "confirmed"

            # Partner
            partner = None
            if "(ob)" in t_low or "orange" in t_low or " ob " in f" {t_low} ":
                partner = "OBS"
            elif "att" in t_low or "at&t" in t_low:
                partner = "AT&T"
            elif "(saf)" in t_low or "safran" in t_low:
                partner = "Safran"

            # Fuzzy site_id match (only for non-spare)
            matched = None
            if confidence != "spare":
                cleaned = _clean_display_name(text)
                key = norm_key(cleaned)
                # Check manual alias first
                for alias_key, sid in WEEKLY_PLAN_ALIASES.items():
                    if alias_key in key and sid in sites:
                        matched = sid
                        break
                if matched:
                    rows.append({
                        "week_label": label,
                        "column_position": j,
                        "row_position": i,
                        "display_name": text,
                        "site_id": matched,
                        "confidence": confidence,
                        "partner": partner,
                    })
                    continue
                # Longest common prefix match (at least 6 chars)
                best_score = 0
                for k in sorted_keys:
                    if len(k) < 6:
                        continue
                    # Full substring match preferred
                    if k in key:
                        matched = site_norm_index[k]
                        best_score = len(k) + 1000
                        break
                    # Fall back to prefix match
                    prefix = 0
                    for a, b in zip(k, key):
                        if a == b:
                            prefix += 1
                        else:
                            break
                    if prefix >= 6 and prefix > best_score:
                        matched = site_norm_index[k]
                        best_score = prefix

            rows.append({
                "week_label": label,
                "column_position": j,
                "row_position": i,
                "display_name": text,
                "site_id": matched,
                "confidence": confidence,
                "partner": partner,
            })
    return rows


def extract_risk_history(ws):
    """Rows 2-7 = aggregate trend. Cols 2=current, 5=WK13, 6=WK14, 7=WK15, 8=WK16."""
    week_cols = [("current", 2), ("WK13", 5), ("WK14", 6), ("WK15", 7), ("WK16", 8)]
    bucket_rows = {
        "sites_migrated": 2,
        "no_risk": 3,
        "low_risk": 4,
        "medium_risk": 5,
        "high_risk": 6,
        "not_possible_to_plan": 7,
    }
    scope_row = 8  # "number of sites by July (219)"

    rows = []
    scope_by_week = {}
    for wk, col in week_cols:
        v = ws.cell(row=scope_row, column=col).value
        if isinstance(v, (int, float)):
            scope_by_week[wk] = int(v)

    for bucket, row in bucket_rows.items():
        for wk, col in week_cols:
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)):
                rows.append({
                    "week_iso": wk,
                    "bucket": bucket,
                    "count": int(v),
                    "total_scope": scope_by_week.get(wk),
                })
    return rows


def extract_kpis(ws):
    """KPIs sheet: warmup % by company + to-be-migrated by site type."""
    rows = []
    # Left pivot: "Nombre de Status" rows 4-7
    for i in range(4, 8):
        label = ws.cell(row=i, column=2).value
        count = ws.cell(row=i, column=3).value
        if label and isinstance(count, (int, float)):
            rows.append({
                "period": "current",
                "metric": f"to_be_migrated_{str(label).lower().replace(' ', '_').replace('-', '_')}",
                "value": int(count),
                "target": None,
                "unit": "sites",
                "comment": f"Source: KPIs sheet pivot (row {i})",
            })
    # Right pivot: warmup % by company rows 4-18 (col 17=company, 18=nb, 19=pct)
    for i in range(4, 19):
        company = ws.cell(row=i, column=17).value
        nb = ws.cell(row=i, column=18).value
        pct = ws.cell(row=i, column=19).value
        if company and isinstance(pct, (int, float)):
            rows.append({
                "period": "current",
                "metric": f"warmup_pct_{str(company).strip().replace(' ', '_').lower()}",
                "value": round(float(pct) * 100, 1),
                "target": 100.0,
                "unit": "pct",
                "comment": f"{int(nb) if isinstance(nb, (int, float)) else ''} sites",
            })
    return rows


def build_decom_backlog(sites):
    """One row per migrated site, with decom status."""
    rows = []
    for sid, rec in sites.items():
        if rec["status"] not in ("migrated", "closed"):
            continue
        mig = rec["migration_date"]
        days_since = None
        if mig:
            try:
                md = datetime.strptime(mig, "%Y-%m-%d").date()
                days_since = (TODAY - md).days
            except ValueError:
                pass
        decom_done = bool(rec["decom_change"]) and rec["decom_change"] not in ("TBC", "TO DO")
        # Overdue: migrated > 60 days ago and no decom change
        overdue = (days_since is not None and days_since > 60 and not decom_done)
        rows.append({
            "site_id": sid,
            "company": rec["company"],
            "migration_date": mig,
            "decom_change": rec["decom_change"],
            "decom_status": rec["decom_status"],
            "decom_date": rec["decom_date"],
            "decom_done": decom_done,
            "days_since_migration": days_since,
            "overdue": overdue,
        })
    return rows


def finalize_sites(sites):
    """Apply final classification and cleanup."""
    for sid, rec in sites.items():
        status, inferred = classify_status(
            rec["status_raw"],
            rec["migration_date"],
            rec["remediation_open"],
        )
        rec["status"] = status
        rec["status_inferred"] = inferred
        # Clean: convert everything to str/number/None, flatten source_sheets to comma list
        rec["source_sheets"] = ", ".join(rec["source_sheets"])
        # Compute t_minus_days if still missing and we have a migration date
        if rec["t_minus_days"] is None and rec["migration_date"]:
            try:
                md = datetime.strptime(rec["migration_date"], "%Y-%m-%d").date()
                rec["t_minus_days"] = (md - TODAY).days
            except (ValueError, TypeError):
                pass


def write_xlsx(path, sites_list, prereqs_long, weekly_plan, risk_history,
               remediation, decom_backlog, kpis, orphans_log, prereq_meta, dupes_log, invalid_log,
               intern_path, sprint_path):
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ---------- 0_README ----------
    ws = wb.create_sheet("0_README")
    ws.append(["sdwan_dashboard_input.xlsx"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Extraction date", TODAY.strftime("%Y-%m-%d")])
    ws.append(["As-of date", TODAY.strftime("%Y-%m-%d")])
    ws.append(["Source file 1", str(intern_path.name)])
    ws.append(["Source file 2", str(sprint_path.name)])
    ws.append([])
    ws.append(["Sheet", "Grain", "Row count"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.cell(row=ws.max_row, column=2).font = bold
    ws.cell(row=ws.max_row, column=3).font = bold
    ws.append(["sites", "1 row per site (canonical scope = UNION(BATCH, Sites list))", len(sites_list)])
    ws.append(["prereqs", "1 row per (site × prereq)", len(prereqs_long)])
    ws.append(["weekly_plan", "1 row per weekly slot", len(weekly_plan)])
    ws.append(["risk_history", "1 row per (week × bucket)", len(risk_history)])
    ws.append(["remediation", "1 row per open ticket", len(remediation)])
    ws.append(["decom_backlog", "1 row per migrated site", len(decom_backlog)])
    ws.append(["kpis", "1 row per (period × metric)", len(kpis)])
    ws.append([])
    ws.append(["Normalization rules:"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for rule in NORMALIZATION_RULES:
        ws.append([rule])
    ws.append([])
    ws.append(["Data quality warnings:"])
    ws.cell(row=ws.max_row, column=1).font = bold
    if dupes_log:
        for site, r1, r2 in dupes_log:
            ws.append([f"Duplicate Site ID in Sites list", site, f"rows {r1} and {r2} — kept first occurrence"])
    if invalid_log:
        ws.append([])
        ws.append([f"Invalid Site IDs filtered out ({len(invalid_log)} entries):"])
        for source, row, val in invalid_log:
            ws.append([source, f"row {row}", val])
    ws.append([])
    ws.append(["Orphans (unresolved references, by source sheet):"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for source, entries in orphans_log.items():
        uniq = sorted(set(entries))
        ws.append([source, len(uniq), "; ".join(uniq[:20]) + (" ..." if len(uniq) > 20 else "")])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 90

    # ---------- sites ----------
    ws = wb.create_sheet("sites")
    cols = [
        "site_id", "company", "business_unit", "country", "region",
        "topology", "batch_year",
        "status", "status_raw", "status_inferred", "partial_migration",
        "migration_date", "migration_date_baseline", "t_minus_days",
        "go_nogo_date", "chrono_review_date", "hypercare",
        "underlay_asis", "underlay_target", "bandwidth",
        "prereq_progress_pct",
        "risk_level", "risk_underlay", "risk_overlay", "risk_safran", "risk_other",
        "remediation_open",
        "decom_change", "decom_status", "decom_date",
        "referent", "coordinator",
        "address", "it_contact", "facilities_contact",
        "comment", "source_sheets",
    ]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for rec in sites_list:
        ws.append([rec.get(c) for c in cols])
    ws.freeze_panes = "A2"
    for col_letter in ["A", "B", "H"]:
        ws.column_dimensions[col_letter].width = 20

    # ---------- prereqs ----------
    ws = wb.create_sheet("prereqs")
    cols = ["site_id", "prereq_key", "when_days", "when_label", "owner",
            "value_pct", "value_raw", "due_date", "overdue"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for r in prereqs_long:
        ws.append([r.get(c) for c in cols])
    ws.freeze_panes = "A2"

    # ---------- weekly_plan ----------
    ws = wb.create_sheet("weekly_plan")
    cols = ["week_label", "column_position", "row_position",
            "display_name", "site_id", "confidence", "partner"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for r in weekly_plan:
        ws.append([r.get(c) for c in cols])
    ws.freeze_panes = "A2"
    ws.column_dimensions["D"].width = 50

    # ---------- risk_history ----------
    ws = wb.create_sheet("risk_history")
    cols = ["week_iso", "bucket", "count", "total_scope"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for r in risk_history:
        ws.append([r.get(c) for c in cols])
    ws.freeze_panes = "A2"

    # ---------- remediation ----------
    ws = wb.create_sheet("remediation")
    cols = ["display_name", "site_id", "category", "severity",
            "root_cause", "blocker", "action_log"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for r in remediation:
        ws.append([r.get(c) for c in cols])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["G"].width = 80

    # ---------- decom_backlog ----------
    ws = wb.create_sheet("decom_backlog")
    cols = ["site_id", "company", "migration_date", "decom_change", "decom_status",
            "decom_date", "decom_done", "days_since_migration", "overdue"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for r in decom_backlog:
        ws.append([r.get(c) for c in cols])
    ws.freeze_panes = "A2"

    # ---------- kpis ----------
    ws = wb.create_sheet("kpis")
    cols = ["period", "metric", "value", "target", "unit", "comment"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).fill = header_fill
    for r in kpis:
        ws.append([r.get(c) for c in cols])
    ws.freeze_panes = "A2"
    ws.column_dimensions["B"].width = 40

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Build clean dashboard input from raw PMO xlsx files.")
    parser.add_argument("--intern", default=str(DEFAULT_INTERN),
                        help="Path to SDWAN_Timeline_Intern*.xlsx (default: latest match in project root)")
    parser.add_argument("--sprint", default=str(DEFAULT_SPRINT),
                        help="Path to Sprint SD-WAN Week - RAMPUP*.xlsx (default: latest match)")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT),
                        help="Output xlsx path (default: sdwan_dashboard_input.xlsx in project root)")
    parser.add_argument("--as-of", default=None,
                        help="As-of date (YYYY-MM-DD). Default: today.")
    args = parser.parse_args()

    global TODAY
    if args.as_of:
        TODAY = datetime.strptime(args.as_of, "%Y-%m-%d").date()

    intern_path = Path(args.intern)
    sprint_path = Path(args.sprint)
    output_path = Path(args.output)

    print(f"Reading {intern_path.name}")
    wb_i = openpyxl.load_workbook(intern_path, data_only=True)
    print(f"Reading {sprint_path.name}")
    wb_s = openpyxl.load_workbook(sprint_path, data_only=True)
    print(f"As-of date: {TODAY}")

    orphans_log = defaultdict(list)

    # Canonical scope = UNION(BATCH 25-26, PILOTS) with aliases + exclusions
    dupes_log = []
    invalid_log = []
    excluded_log = []
    sites, by_canon = load_canonical_sites(wb_i, dupes_log, invalid_log, excluded_log)
    resolve = make_resolver(by_canon, orphans_log)

    # Enrich
    prereq_meta, prereqs_long = enrich_from_batch(sites, resolve, wb_i["BATCH 25-26"])
    enrich_from_risk(sites, resolve, wb_s["Risk follow-up"])
    enrich_from_miseenforme(sites, resolve, wb_i["Mise en forme"])
    enrich_from_decomm(sites, resolve, wb_i["Decomm + SNOW + Dynatrace"])

    # Orange Business operational tracker (parallel "OB delivery" view)
    print("Extracting Orange operational pre-requisites …")
    ob_prereqs, ob_meta, ob_stats = extract_orange_prereqs(resolve, TODAY)
    if ob_stats.get("sdwan_file"):
        print(f"  → source: {ob_stats['sdwan_file']}")
    if ob_stats.get("underlay_file"):
        print(f"  → underlay: {ob_stats['underlay_file']}")
    print(f"  → {ob_stats['matched']} sites matched, {ob_stats['orphans']} orphans, "
          f"{ob_stats['prereqs_emitted']} prereq rows")
    prereqs_long.extend(ob_prereqs)
    prereq_meta.update(ob_meta)
    for canon_id in {p["site_id"] for p in ob_prereqs}:
        if canon_id in sites and "Orange tracker" not in sites[canon_id]["source_sheets"]:
            sites[canon_id]["source_sheets"].append("Orange tracker")

    # Derived tables
    remediation = extract_remediation(wb_s["Remédiation"], sites)
    weekly_plan = extract_weekly_plan(wb_s["Planif migrations per week"], sites)
    risk_history = extract_risk_history(wb_s["Risk follow-up"])
    kpis = extract_kpis(wb_i["KPIs"])

    # Final classification (uses remediation_open flag)
    finalize_sites(sites)

    decom_backlog = build_decom_backlog(sites)

    # Stable order: sites alphabetical, within site Safran prereqs first (in PREREQ_COLS
    # order), then OB prereqs (in OB_PREREQS order). Unknown keys go last.
    prereq_order = list(PREREQ_COLS.keys()) + [k for k, *_ in OB_PREREQS]
    order_idx = {k: i for i, k in enumerate(prereq_order)}
    prereqs_long.sort(
        key=lambda r: (r["site_id"], order_idx.get(r["prereq_key"], len(order_idx)))
    )

    # Sort sites
    sites_list = sorted(sites.values(), key=lambda r: r["site_id"])

    # Sort weekly plan by column order (time), then row
    weekly_plan.sort(key=lambda r: (r["column_position"], r["row_position"]))

    # Write
    write_xlsx(output_path, sites_list, prereqs_long, weekly_plan, risk_history,
               remediation, decom_backlog, kpis, orphans_log, prereq_meta, dupes_log, invalid_log,
               intern_path, sprint_path)
    print(f"\nWrote {output_path}")

    # --- Summary ---
    status_cnt = Counter(r["status"] for r in sites.values())
    print(f"\n=== sites ({len(sites_list)}) ===")
    for s in ("closed", "migrated", "blocked", "ready", "planned", "postponed", "unknown"):
        print(f"  {s:<12} {status_cnt.get(s, 0):>4}")

    print(f"\n=== prereqs ({len(prereqs_long)}) ===")
    overdue_ct = sum(1 for r in prereqs_long if r["overdue"])
    print(f"  overdue: {overdue_ct}")

    print(f"\n=== weekly_plan ({len(weekly_plan)}) ===")
    conf = Counter(r["confidence"] for r in weekly_plan)
    print(f"  {dict(conf)}")
    unresolved = sum(1 for r in weekly_plan if r["site_id"] is None)
    print(f"  unresolved site_ids: {unresolved} / {len(weekly_plan)}")

    print(f"\n=== risk_history: {len(risk_history)} rows")
    print(f"=== remediation ({len(remediation)}) ===")
    sev = Counter(r["severity"] for r in remediation)
    matched_rem = sum(1 for r in remediation if r["site_id"])
    print(f"  {dict(sev)}  | matched to site_id: {matched_rem}/{len(remediation)}")

    print(f"\n=== decom_backlog ({len(decom_backlog)}) ===")
    overdue_decom = sum(1 for r in decom_backlog if r["overdue"])
    print(f"  overdue (>60j sans CHG): {overdue_decom}")

    print(f"\n=== kpis: {len(kpis)} rows")

    print(f"\n=== orphans log ===")
    for source, entries in orphans_log.items():
        uniq = sorted(set(entries))
        print(f"  {source}: {len(uniq)} unique")
        for e in uniq[:10]:
            print(f"    - {e}")


if __name__ == "__main__":
    main()
