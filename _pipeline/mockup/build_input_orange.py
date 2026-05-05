"""
Extract Orange Business operational pre-requisites from the two Orange tracker
files and emit them as prereq rows compatible with build_input.py's `prereqs_long`
schema (one row per site × prereq).

Source files (latest dated `Onglet suivi *Orange*.xlsx` in project root):
  - "Onglet suivi SDWAN Orange <DDMMYYYY>.xlsx"  → 1 row per site, 80 columns
  - "Onglet suivi underlay Orange <DDMMYYYY>.xlsx" → 130 sites in coordination pool

Output format (returned to build_input.py): same dict shape as prereqs_long rows
plus a prereq_meta dict.

This module is imported by build_input.py — it does not run on its own.
"""

from __future__ import annotations

import re
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import openpyxl


ROOT = Path(__file__).resolve().parent.parent


# ============================================================
# Configuration: which Orange jalons → which prereq keys
# ============================================================
#
# Order matters: this is the chronological order in which the OB delivery
# milestones happen, T0 = migration date. when_days are heuristic offsets
# used by the dashboard for "overdue" detection.

OB_PREREQS = [
    # (key,                       when_days, sdwan_col_idx, mapper)
    ("OB Site Survey",            -84,  17, "site_survey"),
    ("OB BDC Validated",          -77,  32, "bdc"),
    ("OB Equipment On Site",      -42,  40, "equipment"),    # uses col 40 (AN) + 36 (AJ)
    ("OB Underlay RFS",           -28,  43, "yesno"),         # AQ "Underlay ready"
    ("OB Site Installed",         -14,  52, "yesno"),         # AZ "Site installed"
    ("OB Migration Chronogram",    -7,  56, "yesno"),         # BD "Migration chronogram Ready"
    ("OB RFS Closure",              0,  60, "closure"),       # BH "Closure Status (RFS/RFB)"
    ("OB SI Closed",               56,  63, "si_closed"),     # BK "SI closed (Fully Migrated)"
]


# ============================================================
# Mappers — convert raw Orange status text to 0/50/100 (or None)
# ============================================================

def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def map_site_survey(v, _row=None) -> Optional[float]:
    s = _norm(v)
    if s == "completed":  return 100.0
    if s == "scheduled":  return 50.0
    if s == "proposed":   return 25.0
    if s == "pending":    return 25.0
    if s == "n/a":        return None
    if s == "":           return 0.0
    return 0.0


def map_bdc(v, _row=None) -> Optional[float]:
    s = _norm(v)
    if s == "validated":         return 100.0
    if "pending ob" in s:        return 50.0
    if "pending safran" in s:    return 25.0
    if s == "":                  return 0.0
    return 0.0


def map_equipment(v, ctx=None) -> Optional[float]:
    """Use col AN 'Equipment available' (Yes/No) primarily; col AJ 'Equipment Status'
    refines the partial states."""
    avail = _norm(v)
    eq_status = _norm(ctx) if ctx is not None else ""
    if avail == "yes":           return 100.0
    if eq_status == "on site":   return 100.0
    if eq_status == "at warehouse": return 50.0
    if eq_status == "on-going":  return 50.0
    if eq_status == "ordered":   return 25.0
    if eq_status == "to be ordered": return 0.0
    if eq_status == "n/a":       return None
    if avail == "no":            return 0.0
    return 0.0


def map_yesno(v, _row=None) -> Optional[float]:
    s = _norm(v)
    if s == "yes":           return 100.0
    if s == "no":            return 0.0
    if "cancelled" in s or "annulé" in s: return None
    if s == "n/a":           return None
    return 0.0


def map_closure(v, _row=None) -> Optional[float]:
    s = _norm(v)
    if s == "successful":    return 100.0
    if s == "failed":        return 0.0
    if s == "":              return 0.0
    return 0.0


def map_si_closed(v, _row=None) -> Optional[float]:
    s = _norm(v)
    if s == "closed":        return 100.0
    if s == "in progress":   return 50.0
    if s == "":              return 0.0
    return 0.0


MAPPERS = {
    "site_survey": map_site_survey,
    "bdc":         map_bdc,
    "equipment":   map_equipment,
    "yesno":       map_yesno,
    "closure":     map_closure,
    "si_closed":   map_si_closed,
}


# ============================================================
# File discovery
# ============================================================

def _latest_orange_file(name_root: str) -> Optional[Path]:
    """Find the most recent `Onglet suivi <name_root> Orange*.xlsx` by embedded date."""
    pattern = f"Onglet suivi {name_root} Orange*.xlsx"
    candidates = [c for c in ROOT.glob(pattern) if not c.name.startswith("~$")]
    if not candidates:
        return None
    dated = []
    for c in candidates:
        m = re.search(r"(\d{2})(\d{2})(\d{4})", c.name)
        if m:
            dd, mm, yyyy = m.groups()
            dated.append((f"{yyyy}{mm}{dd}", c))
    if dated:
        dated.sort(key=lambda x: x[0], reverse=True)
        return dated[0][1]
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def find_orange_files() -> tuple[Optional[Path], Optional[Path]]:
    return _latest_orange_file("SDWAN"), _latest_orange_file("underlay")


# ============================================================
# Extractor
# ============================================================

def extract_orange_prereqs(resolve, today: date):
    """Read Orange tracker files and return:
        prereqs_long: list of dicts compatible with build_input.py
        prereq_meta:  dict {key → {when_days, when_label, owner}}
        stats:        dict with extraction counters

    `resolve(raw_site_id, source_label)` is the canonical site resolver from
    build_input.py — returns canonical site_id or None (logs as orphan).
    """
    sdwan_path, underlay_path = find_orange_files()
    stats = {
        "sdwan_file": sdwan_path.name if sdwan_path else None,
        "underlay_file": underlay_path.name if underlay_path else None,
        "sdwan_rows": 0,
        "underlay_rows": 0,
        "matched": 0,
        "orphans": 0,
        "prereqs_emitted": 0,
    }

    if sdwan_path is None:
        print("  ⚠ No 'Onglet suivi SDWAN Orange*.xlsx' found — skipping OB extraction.")
        return [], {}, stats

    wb = openpyxl.load_workbook(sdwan_path, data_only=True)
    if "Safran Roll-out" not in wb.sheetnames:
        print(f"  ⚠ Sheet 'Safran Roll-out' missing in {sdwan_path.name} — skipping.")
        return [], {}, stats
    ws = wb["Safran Roll-out"]

    # ----- Underlay RFS dates (optional enrichment) -----
    rfs_dates = {}  # canonical site_id → RFS datetime
    if underlay_path is not None:
        wb_u = openpyxl.load_workbook(underlay_path, data_only=True)
        sheet_name = next((s for s in wb_u.sheetnames if "Pool site" in s), None)
        if sheet_name:
            ws_u = wb_u[sheet_name]
            for i in range(3, ws_u.max_row + 1):
                stats["underlay_rows"] += 1
                raw = ws_u.cell(row=i, column=2).value
                if not raw:
                    continue
                # Strip parenthesized qualifiers like "(FTTO)" / "(FTTE)" so we
                # can match the canonical site
                clean = re.sub(r"\s*\([^)]*\)\s*", "", str(raw)).strip()
                canon = resolve(clean, "Onglet suivi underlay Orange")
                if canon is None:
                    continue
                rfs = ws_u.cell(row=i, column=8).value
                if isinstance(rfs, (datetime, date)):
                    d = rfs.date() if isinstance(rfs, datetime) else rfs
                    if canon not in rfs_dates or d < rfs_dates[canon]:
                        rfs_dates[canon] = d

    # ----- Build prereq_meta -----
    prereq_meta = {}
    for key, when_days, _col, _mapper in OB_PREREQS:
        prereq_meta[key] = {
            "when_days": when_days,
            "when_label": (
                "T0" if when_days == 0
                else (f"T{when_days // 7}W" if when_days % 7 == 0 else f"T{when_days}D")
            ),
            "owner": "OBS",
            "owner_raw": "Orange Business (operational tracker)",
        }

    # ----- Walk SDWAN tracker rows -----
    prereqs_long = []
    matched_sites = set()
    orphan_count = 0

    for r in range(2, ws.max_row + 1):
        site_raw = ws.cell(row=r, column=1).value
        if not site_raw:
            continue
        stats["sdwan_rows"] += 1
        canon = resolve(str(site_raw).strip(), "Onglet suivi SDWAN Orange")
        if canon is None:
            orphan_count += 1
            continue
        matched_sites.add(canon)

        # Migration date for due_date computation: prefer planned (col 57 BE),
        # fall back to target (col 58 BF).
        mig_raw = ws.cell(row=r, column=57).value or ws.cell(row=r, column=58).value
        mig_date = None
        if isinstance(mig_raw, (datetime, date)):
            mig_date = mig_raw.date() if isinstance(mig_raw, datetime) else mig_raw

        for key, when_days, col, mapper_name in OB_PREREQS:
            cell = ws.cell(row=r, column=col)
            mapper = MAPPERS[mapper_name]
            ctx = ws.cell(row=r, column=36).value if mapper_name == "equipment" else None
            value_pct = mapper(cell.value, ctx)
            value_raw = str(cell.value).strip() if cell.value is not None else None

            due_date = None
            overdue = None
            if mig_date is not None:
                due = date.fromordinal(mig_date.toordinal() + when_days)
                due_date = due.strftime("%Y-%m-%d")
                overdue = (today >= due) and (value_pct is not None and value_pct < 100)

            prereqs_long.append({
                "site_id": canon,
                "prereq_key": key,
                "when_days": when_days,
                "when_label": prereq_meta[key]["when_label"],
                "owner": "OBS",
                "value_pct": value_pct,
                "value_raw": value_raw,
                "due_date": due_date,
                "overdue": overdue,
            })

    stats["matched"] = len(matched_sites)
    stats["orphans"] = orphan_count
    stats["prereqs_emitted"] = len(prereqs_long)
    stats["rfs_dates_collected"] = len(rfs_dates)

    return prereqs_long, prereq_meta, stats
